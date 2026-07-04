"""One-time importer: Kitchen Store & Consumption Log.xlsx -> MongoDB.

Runs on the HOST (needs openpyxl + pymongo), connecting to the mapped Mongo port.
It is idempotent for masters (upsert by unique name) and replaces staging data.

P0 scope:
  * Masters: units, item_categories, items, vendors, canteens, staff, consumers.
  * Staging (raw history for later phases): legacy_purchases, legacy_inward,
    legacy_outward, legacy_consumption.

It does NOT create stock movements / cost layers / invoices — those are produced
by the P1/P2 store & procurement engines once they exist.

Usage (from project root):
  python backend/scripts/import_legacy.py \
      --xlsx "Kitchen Store & Consumption Log.xlsx" \
      --mongo "mongodb://localhost:27019/?directConnection=true" \
      --db sspl_kitchen
"""
from __future__ import annotations

import argparse
import datetime as dt
import io
import sys
from typing import Any

import openpyxl
from pymongo import MongoClient, UpdateOne

# Make stdout UTF-8 safe (item names contain non-cp1252 characters on Windows).
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

EXCEL_EPOCH = dt.datetime(1899, 12, 30)  # Excel serial-date origin


# ---------- value normalisation ----------
def norm_date(v: Any) -> dt.datetime | None:
    """Normalise Excel dates: datetime, dd-mm-yyyy strings, or serial floats."""
    if v is None or v == "":
        return None
    if isinstance(v, dt.datetime):
        return v
    if isinstance(v, dt.date):
        return dt.datetime(v.year, v.month, v.day)
    if isinstance(v, (int, float)):
        try:
            return EXCEL_EPOCH + dt.timedelta(days=float(v))
        except (ValueError, OverflowError):
            return None
    s = str(v).strip()
    for fmt in ("%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return dt.datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def clean(v: Any) -> str:
    return "" if v is None else str(v).strip()


def to_float(v: Any) -> float | None:
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(",", "").strip())
    except ValueError:
        return None


# ---------- classifiers ----------
def classify_category(name: str) -> str:
    n = name.lower()
    if any(k in n for k in ("gas", "coal", "lpg", "cylinder")):
        return "Gas & Fuel"
    if any(k in n for k in ("oil", "ghee", "dalda")):
        return "Oil & Ghee"
    return "Grocery"


def classify_consumer(name: str) -> str:
    n = name.lower()
    if "company staff" in n or "company helper" in n:
        return "company_group"
    if any(k in n for k in ("security", "mechanic", "air pipe", "lift", "pipe")):
        return "service"
    return "contractor"


# ---------- master upsert helpers ----------
def upsert_named(coll, names: list[str], extra_by_name: dict | None = None) -> dict[str, Any]:
    """Upsert simple named docs; return {name: _id}."""
    ops = []
    for name in names:
        doc = {"name": name, "active": True}
        if extra_by_name and name in extra_by_name:
            doc.update(extra_by_name[name])
        ops.append(UpdateOne({"name": name}, {"$setOnInsert": doc}, upsert=True))
    if ops:
        coll.bulk_write(ops, ordered=False)
    return {d["name"]: d["_id"] for d in coll.find({}, {"name": 1})}


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", default="Kitchen Store & Consumption Log.xlsx")
    ap.add_argument("--mongo", default="mongodb://localhost:27019/?directConnection=true")
    ap.add_argument("--db", default="sspl_kitchen")
    args = ap.parse_args()

    print(f"Reading {args.xlsx} ...")
    wb = openpyxl.load_workbook(args.xlsx, data_only=True, read_only=True)
    client = MongoClient(args.mongo)
    db = client[args.db]

    def rows(sheet: str, start: int):
        ws = wb[sheet]
        for r in ws.iter_rows(min_row=start, values_only=True):
            if r and any(c not in (None, "") for c in r):
                yield r

    # ---- 1. Units (from Stock Data col D) + Items (col C) ----
    stock_products: list[tuple[str, str]] = []  # (product, unit)
    units_set: set[str] = set()
    for r in rows("Stock Data", 2):
        product, unit = clean(r[2] if len(r) > 2 else ""), clean(r[3] if len(r) > 3 else "")
        if product:
            stock_products.append((product, unit))
            if unit:
                units_set.add(unit)
    unit_ids = upsert_named(db.units, sorted(units_set))
    print(f"  units: {len(unit_ids)}")

    # ---- 2. Categories (keyword-derived) ----
    cat_names = sorted({classify_category(p) for p, _ in stock_products})
    cat_ids = upsert_named(db.item_categories, cat_names)
    print(f"  categories: {list(cat_ids)}")

    # ---- 3. Items (66 from Stock Data) ----
    item_ops = []
    for product, unit in stock_products:
        item_ops.append(
            UpdateOne(
                {"name": product, "category_id": cat_ids[classify_category(product)]},
                {
                    "$setOnInsert": {
                        "name": product,
                        "category_id": cat_ids[classify_category(product)],
                        "unit_id": unit_ids.get(unit),
                        "min_stock": None,
                        "max_stock": None,
                        "active": True,
                    }
                },
                upsert=True,
            )
        )
    if item_ops:
        db.items.bulk_write(item_ops, ordered=False)
    item_by_name = {d["name"].lower(): d["_id"] for d in db.items.find({}, {"name": 1})}
    print(f"  items: {len(item_by_name)}")

    # ---- 4. Vendors (distinct Party Name from Kitchen Material Inward, col F idx5) ----
    vendors_set: set[str] = set()
    for r in rows("Kitchen Material Inward", 3):
        party = clean(r[5] if len(r) > 5 else "")
        if party:
            vendors_set.add(party)
    vendor_ids = upsert_named(db.vendors, sorted(vendors_set))
    print(f"  vendors: {len(vendor_ids)}")

    # ---- 5. Canteens + Staff (Canteen Name Sheet) ----
    canteen_set: set[str] = set()
    staff_rows: list[tuple[str, str, str, str, str]] = []
    for r in rows("Canteen Name Sheet", 2):
        canteen = clean(r[0] if len(r) > 0 else "")
        name = clean(r[1] if len(r) > 1 else "")
        occ = clean(r[2] if len(r) > 2 else "")
        time = clean(r[3] if len(r) > 3 else "")
        rest = clean(r[4] if len(r) > 4 else "")
        if canteen:
            canteen_set.add(canteen)
        if canteen and name:
            staff_rows.append((canteen, name, occ, time, rest))
    canteen_ids = upsert_named(db.canteens, sorted(canteen_set))
    db.staff.delete_many({})  # roster is small & fully re-derivable
    if staff_rows:
        db.staff.insert_many(
            [
                {
                    "name": name,
                    "canteen_id": canteen_ids[canteen],
                    "occupation": occ or None,
                    "shift_time": time or None,
                    "rest_time": rest or None,
                    "active": True,
                }
                for canteen, name, occ, time, rest in staff_rows
            ]
        )
    print(f"  canteens: {len(canteen_ids)}  staff: {len(staff_rows)}")

    # ---- 6. Consumers (50 entities from Mess Daily Food Con header row 2, cols C+) ----
    mess = wb["Mess Daily Food Con."]
    header = next(mess.iter_rows(min_row=2, max_row=2, values_only=True))
    consumer_names = [clean(h) for h in header[2:] if clean(h)]
    consumer_extra = {
        n: {"type": classify_consumer(n), "billable": False, "meal_rate": None}
        for n in consumer_names
    }
    consumer_ids = upsert_named(db.consumers, consumer_names, consumer_extra)
    print(f"  consumers: {len(consumer_ids)}")

    # ---- 7. Stage raw transaction history ----
    # 7a. legacy_purchases (Kitchen Material Inward)
    db.legacy_purchases.delete_many({})
    purchases = []
    for r in rows("Kitchen Material Inward", 3):
        g = lambda i: r[i] if len(r) > i else None  # noqa: E731
        if not clean(g(6)):  # no item name
            continue
        purchases.append(
            {
                "timestamp": norm_date(g(0)),
                "email": clean(g(1)) or None,
                "invoice_date": norm_date(g(2)),
                "invoice_number": clean(g(3)) or None,
                "invoice_photo": clean(g(4)) or None,
                "party_name": clean(g(5)) or None,
                "item_name": clean(g(6)),
                "quantity": to_float(g(7)),
                "rate": to_float(g(8)),
                "unit": clean(g(9)) or None,
                "total": to_float(g(10)),
                "remarks": clean(g(11)) or None,
            }
        )
    if purchases:
        db.legacy_purchases.insert_many(purchases)

    # 7b. legacy_inward (Inward Data)
    db.legacy_inward.delete_many({})
    inward = []
    for r in rows("Inward Data", 2):
        g = lambda i: r[i] if len(r) > i else None  # noqa: E731
        if not clean(g(2)):
            continue
        inward.append(
            {
                "date": norm_date(g(0)),
                "received_by": clean(g(1)) or None,
                "item_name": clean(g(2)),
                "unit": clean(g(3)) or None,
                "quantity": to_float(g(4)),
                "submitted_by": clean(g(5)) or None,
                "remark": clean(g(6)) or None,
            }
        )
    if inward:
        db.legacy_inward.insert_many(inward)

    # 7c. legacy_outward (Outward Data) — date in col0, issuer col1
    db.legacy_outward.delete_many({})
    outward = []
    for r in rows("Outward Data", 2):
        g = lambda i: r[i] if len(r) > i else None  # noqa: E731
        if not clean(g(2)):
            continue
        outward.append(
            {
                "date": norm_date(g(0)),
                "issued_by": clean(g(1)) or None,
                "item_name": clean(g(2)),
                "unit": clean(g(3)) or None,
                "quantity": to_float(g(4)),
                "submitted_by": clean(g(5)) or None,
                "remark": clean(g(6)) or None,
            }
        )
    if outward:
        db.legacy_outward.insert_many(outward)

    # 7d. legacy_consumption (Mess Daily Food Con) -> one row per (date,shift,consumer)
    db.legacy_consumption.delete_many({})
    consumption = []
    for r in mess.iter_rows(min_row=3, values_only=True):
        d = norm_date(r[0] if len(r) > 0 else None)
        shift = clean(r[1] if len(r) > 1 else "")
        if d is None or not shift:
            continue
        for idx, cname in enumerate(consumer_names, start=2):
            cnt = to_float(r[idx]) if len(r) > idx else None
            if cnt:
                consumption.append(
                    {"date": d, "shift": shift, "consumer_name": cname, "count": int(cnt)}
                )
    if consumption:
        db.legacy_consumption.insert_many(consumption)

    print(
        f"  staged -> purchases:{len(purchases)} inward:{len(inward)} "
        f"outward:{len(outward)} consumption:{len(consumption)}"
    )

    # ---- 8. Report item-name mismatches (transactions referencing unknown items) ----
    txn_item_names = {p["item_name"].lower() for p in purchases}
    txn_item_names |= {i["item_name"].lower() for i in inward}
    txn_item_names |= {o["item_name"].lower() for o in outward}
    unknown = sorted(n for n in txn_item_names if n not in item_by_name)
    print(f"\n  {len(unknown)} transaction item-names not in the 66-item master "
          f"(need alias/merge curation in UI):")
    for n in unknown[:40]:
        print(f"    - {n}")
    if len(unknown) > 40:
        print(f"    ... and {len(unknown) - 40} more")

    client.close()
    print("\nDone.")


if __name__ == "__main__":
    main()
